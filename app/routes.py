from datetime import date, datetime
from functools import wraps
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from flask import Blueprint, abort, flash, jsonify, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required, login_user, logout_user
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.exc import IntegrityError
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from app.config import UPLOADS_DIR
from app.extensions import db
from app.models import Progress, Project, ProjectStudent, Stage, User
from app.service_models import ActionLog, ApiLog, Message


main_bp = Blueprint("main", __name__)

# Подписи статусов для интерфейса.
STATUS_LABELS = {
    "not_started": "не начат",
    "in_progress": "в процессе",
    "done": "выполнен",
    "approved": "проверено",
}
ROLE_LABELS = {
    "teacher": "учитель",
    "student": "ученик",
}
ALLOWED_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}
MAX_MESSAGE_IMAGE_SIZE = 5 * 1024 * 1024
STAGE_IMPORT_REQUIRED_COLUMNS = ("order_num", "title", "description", "deadline")
STAGE_IMPORT_TEMPLATE_HEADERS_RU = (
    "Порядковый номер",
    "Название этапа",
    "Описание этапа",
    "Дедлайн",
)
STAGE_IMPORT_COLUMN_ALIASES = {
    "order_num": ("order_num", "порядковый номер", "номер", "номер этапа"),
    "title": ("title", "название", "название этапа"),
    "description": ("description", "описание", "описание этапа"),
    "deadline": ("deadline", "дедлайн", "срок", "дата дедлайна"),
}
XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLSX_STATUS_FILLS = {
    "success": PatternFill("solid", fgColor="C6EFCE"),
    "warning": PatternFill("solid", fgColor="FFF2CC"),
    "danger": PatternFill("solid", fgColor="F4CCCC"),
    "secondary": PatternFill("solid", fgColor="E7E6E6"),
}
THIN_BORDER = Border(
    left=Side(style="thin", color="B7C6D6"),
    right=Side(style="thin", color="B7C6D6"),
    top=Side(style="thin", color="B7C6D6"),
    bottom=Side(style="thin", color="B7C6D6"),
)

STAGE_ORDER_DUPLICATE_ERROR = "Этап с таким порядковым номером уже есть в проекте."


def role_required(*roles):
    # Ограничивает доступ к маршруту списком ролей.
    def decorator(view):
        # Оборачивает исходный view с проверкой роли.
        @wraps(view)
        def wrapped_view(*args, **kwargs):
            # Прерывает запрос, если роль не входит в разрешенный список.
            if current_user.role not in roles:
                abort(403)
            return view(*args, **kwargs)

        return wrapped_view

    return decorator


def get_teacher_project_or_404(project_id):
    # Возвращает проект, если он принадлежит текущему учителю.
    project = db.session.get(Project, project_id)
    if project is None or project.teacher_id != current_user.id:
        abort(404)
    return project


def get_teacher_stage_or_404(stage_id):
    # Возвращает этап только из проекта текущего учителя.
    stage = db.session.get(Stage, stage_id)
    if stage is None or stage.project.teacher_id != current_user.id:
        abort(404)
    return stage


def get_student_project_or_404(project_id):
    # Возвращает проект, к которому привязан текущий ученик.
    project = db.session.get(Project, project_id)
    if project is None:
        abort(404)

    link = db.session.execute(
        db.select(ProjectStudent).filter_by(project_id=project_id, student_id=current_user.id)
    ).scalar_one_or_none()
    if link is None:
        abort(404)

    return project


def get_student_stage_or_404(stage_id):
    # Проверяет доступ ученика к этапу через проект.
    stage = db.session.get(Stage, stage_id)
    if stage is None:
        abort(404)

    get_student_project_or_404(stage.project_id)
    return stage


def get_project_for_current_user_or_404(project_id):
    # Выбирает проверку доступа к проекту по текущей роли.
    if current_user.role == "teacher":
        return get_teacher_project_or_404(project_id)
    if current_user.role == "student":
        return get_student_project_or_404(project_id)
    abort(403)


def get_or_create_progress(student_id, stage_id):
    # Запись прогресса создается при первом открытии этапа.
    progress = db.session.execute(
        db.select(Progress).filter_by(student_id=student_id, stage_id=stage_id)
    ).scalar_one_or_none()
    if progress is None:
        progress = Progress(student_id=student_id, stage_id=stage_id)
        db.session.add(progress)
        db.session.flush()
    return progress


def log_api_call():
    # Пишет факт вызова API в сервисный журнал.
    api_log = ApiLog(
        user_id=current_user.id if current_user.is_authenticated else None,
        endpoint=request.path,
        method=request.method,
    )
    db.session.add(api_log)


def log_action(action_type, description, project_id=None, user_id=None):
    # Пишет бизнес-событие в журнал действий.
    action_log = ActionLog(
        user_id=user_id if user_id is not None else current_user.id,
        project_id=project_id,
        action_type=action_type,
        description=description,
    )
    db.session.add(action_log)


def get_role_label(role):
    # Возвращает подпись роли для интерфейса.
    return ROLE_LABELS.get(role, role)


def get_status_label(status):
    # Возвращает подпись статуса для интерфейса.
    return STATUS_LABELS.get(status, status)


def get_teacher_unread_by_student(project_id, teacher_id):
    # Считает непрочитанные сообщения учителя по каждому ученику проекта.
    rows = db.session.execute(
        db.select(Message.from_user_id, db.func.count(Message.id))
        .filter(
            Message.project_id == project_id,
            Message.to_user_id == teacher_id,
            Message.is_read.is_(False),
        )
        .group_by(Message.from_user_id)
    ).all()
    return {student_id: unread_count for student_id, unread_count in rows}


def get_current_user_unread_total():
    # Считает общее число непрочитанных сообщений пользователя.
    if not current_user.is_authenticated:
        return 0
    return db.session.execute(
        db.select(db.func.count(Message.id)).filter(
            Message.to_user_id == current_user.id,
            Message.is_read.is_(False),
        )
    ).scalar_one()


def get_nav_messages_href():
    # Колокольчик ведет сразу в последний непрочитанный диалог.
    if not current_user.is_authenticated:
        return None

    unread_rows = db.session.execute(
        db.select(Message.project_id, Message.from_user_id)
        .filter(
            Message.to_user_id == current_user.id,
            Message.is_read.is_(False),
        )
        .order_by(Message.created_at.desc(), Message.id.desc())
    ).all()

    if current_user.role == "teacher":
        for project_id, from_user_id in unread_rows:
            project = db.session.get(Project, project_id)
            if project is None or project.teacher_id != current_user.id:
                continue
            return url_for("main.project_messages", project_id=project_id, student_id=from_user_id)
        return url_for("main.teacher_projects")

    if current_user.role == "student":
        for project_id, from_user_id in unread_rows:
            project = db.session.get(Project, project_id)
            if project is None or project.teacher_id != from_user_id:
                continue
            has_access = db.session.execute(
                db.select(ProjectStudent.id).filter_by(project_id=project_id, student_id=current_user.id)
            ).scalar_one_or_none()
            if has_access is not None:
                return url_for("main.project_messages", project_id=project_id)
        return url_for("main.student_projects")

    return url_for("main.dashboard")


@main_bp.app_context_processor
def inject_nav_unread_counter():
    # Передаем в шапку счетчик непрочитанных и ссылку перехода.
    if not current_user.is_authenticated:
        return {"nav_unread_total": 0, "nav_messages_href": None}
    return {
        "nav_unread_total": get_current_user_unread_total(),
        "nav_messages_href": get_nav_messages_href(),
    }


def validate_and_save_message_image(uploaded_image):
    # Проверяет картинку в сообщении и сохраняет файл в uploads/messages.
    if uploaded_image is None or not uploaded_image.filename:
        return None, None

    safe_name = secure_filename(uploaded_image.filename)
    suffix = Path(safe_name).suffix.lower()
    if suffix not in ALLOWED_IMAGE_EXTENSIONS:
        return None, "Разрешены только изображения: .png, .jpg, .jpeg, .webp."

    uploaded_image.stream.seek(0, 2)
    file_size = uploaded_image.stream.tell()
    uploaded_image.stream.seek(0)
    if file_size > MAX_MESSAGE_IMAGE_SIZE:
        return None, "Размер изображения не должен превышать 5 MB."

    messages_dir = Path(UPLOADS_DIR) / "messages"
    messages_dir.mkdir(parents=True, exist_ok=True)
    target_name = f"{uuid4().hex}{suffix}"
    target_path = messages_dir / target_name
    uploaded_image.save(target_path)
    return str(target_path), None


def serialize_project(project):
    # Преобразует проект в JSON-словарь для API.
    return {
        "id": project.id,
        "title": project.title,
        "description": project.description,
        "teacher_id": project.teacher_id,
        "created_at": project.created_at.isoformat(),
    }


def serialize_stage(stage):
    # Преобразует этап в JSON-словарь для API.
    return {
        "id": stage.id,
        "project_id": stage.project_id,
        "title": stage.title,
        "description": stage.description,
        "deadline": stage.deadline.isoformat(),
        "order_num": stage.order_num,
    }


def style_header_row(worksheet):
    # Оформляет шапку таблицы XLSX.
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def workbook_response(workbook, filename):
    # Отправляет сформированную книгу XLSX в ответ на скачивание.
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype=XLSX_MIMETYPE,
        as_attachment=True,
        download_name=filename,
    )


def build_project_workbook(project):
    # Формирует плоский XLSX-отчет по проекту.
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Отчет"
    worksheet.append(
        [
            "Название проекта",
            "Ученик",
            "Email ученика",
            "Этап",
            "Статус",
            "GitHub ссылка",
            "Комментарий",
            "Обновлено",
        ]
    )

    progress_map = {
        (entry.student_id, entry.stage_id): entry
        for entry in db.session.execute(
            db.select(Progress)
            .join(Stage)
            .filter(Stage.project_id == project.id)
        ).scalars()
    }

    for link in project.student_links:
        for stage in project.stages:
            entry = progress_map.get((link.student_id, stage.id))
            status_value = entry.status if entry else "not_started"
            worksheet.append(
                [
                    project.title,
                    link.student.full_name,
                    link.student.email,
                    stage.title,
                    get_status_label(status_value),
                    entry.github_url if entry else "",
                    entry.comment if entry else "",
                    entry.updated_at if entry and entry.updated_at else "",
                ]
            )

    style_header_row(worksheet)
    column_widths = [24, 24, 28, 28, 16, 36, 42, 22]
    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[7].number_format = "DD.MM.YYYY HH:MM"
    worksheet.freeze_panes = "A2"
    return workbook


def get_stage_status_for_analytics(stage, progress_entry, now):
    # Для незавершенных этапов учитывается просрочка дедлайна.
    status = progress_entry.status if progress_entry else "not_started"
    is_completed = status in {"done", "approved"}
    is_overdue = stage.deadline < now and not is_completed
    if is_overdue:
        return "просрочен"
    return get_status_label(status)


def build_project_analytics_rows(project):
    # Собирает матрицу статусов и процент выполнения по ученикам.
    stages = list(project.stages)
    progress_entries = list(
        db.session.execute(
            db.select(Progress).join(Stage).filter(Stage.project_id == project.id)
        ).scalars()
    )
    progress_map = {(entry.student_id, entry.stage_id): entry for entry in progress_entries}

    student_rows = []
    total_stages = len(stages)
    now = datetime.now()
    for link in project.student_links:
        completed_stages = 0
        stage_cells = []
        for stage in stages:
            entry = progress_map.get((link.student_id, stage.id))
            status = entry.status if entry else "not_started"
            is_completed = status in {"done", "approved"}
            if is_completed:
                completed_stages += 1
            status_label = get_stage_status_for_analytics(stage, entry, now)
            badge_kind = (
                "danger"
                if status_label == "просрочен"
                else "success"
                if is_completed
                else "warning"
                if status == "not_started"
                else "secondary"
            )
            stage_cells.append(
                {
                    "status": status,
                    "status_label": status_label,
                    "badge_kind": badge_kind,
                }
            )

        progress_percent = round((completed_stages / total_stages) * 100) if total_stages else 0
        student_rows.append(
            {
                "student": link.student,
                "completed_stages": completed_stages,
                "total_stages": total_stages,
                "progress_percent": progress_percent,
                "stage_cells": stage_cells,
            }
        )

    return stages, student_rows


def build_project_analytics_workbook(project):
    # Формирует XLSX-аналитику по проекту с цветовой маркировкой статусов.
    stages, student_rows = build_project_analytics_rows(project)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Аналитика"

    header = [
        "Название проекта",
        "Ученик",
        "Email ученика",
    ]
    for stage in stages:
        stage_label = f"{stage.order_num}. {stage.title} ({stage.deadline.strftime('%d.%m.%Y %H:%M')})"
        header.append(stage_label)
    worksheet.append(header)

    for row in student_rows:
        xlsx_row = [
            project.title,
            row["student"].full_name,
            row["student"].email,
        ]
        xlsx_row.extend(cell["status_label"] for cell in row["stage_cells"])
        worksheet.append(xlsx_row)
        current_row = worksheet.max_row
        for offset, cell_data in enumerate(row["stage_cells"], start=4):
            worksheet.cell(row=current_row, column=offset).fill = XLSX_STATUS_FILLS[cell_data["badge_kind"]]

    style_header_row(worksheet)
    worksheet.column_dimensions["A"].width = 26
    worksheet.column_dimensions["B"].width = 26
    worksheet.column_dimensions["C"].width = 30
    for column in range(4, 4 + len(stages)):
        worksheet.column_dimensions[worksheet.cell(row=1, column=column).column_letter].width = 28
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.freeze_panes = "D2"
    return workbook


def parse_deadline_value(raw_value):
    # Поддержка даты из ячейки XLSX и из строкового формата.
    if isinstance(raw_value, datetime):
        return raw_value, None
    if isinstance(raw_value, date):
        return datetime(raw_value.year, raw_value.month, raw_value.day, 0, 0), None

    value = str(raw_value or "").strip()
    if not value:
        return None, "пустой дедлайн"

    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M", "%d.%m.%Y %H:%M"):
        try:
            return datetime.strptime(value, fmt), None
        except ValueError:
            continue
    return None, f"некорректный дедлайн '{value}'"


def parse_order_num_value(raw_value):
    # Номер этапа должен быть целым и больше нуля.
    if isinstance(raw_value, bool):
        return None, "номер этапа должен быть целым числом"
    if isinstance(raw_value, int):
        order_num = raw_value
    elif isinstance(raw_value, float):
        if not raw_value.is_integer():
            return None, "номер этапа должен быть целым числом"
        order_num = int(raw_value)
    else:
        value = str(raw_value or "").strip()
        if not value:
            return None, "пустой номер этапа"
        try:
            order_num = int(value)
        except ValueError:
            return None, f"некорректный номер этапа '{value}'"

    if order_num <= 0:
        return None, "номер этапа должен быть больше нуля"
    return order_num, None


def normalize_stage_import_rows(raw_rows, field_names):
    # Нормализует русские и английские заголовки в единый формат.
    field_map = {}
    for field_name in field_names:
        normalized = str(field_name or "").strip().lower()
        if normalized:
            field_map[normalized] = field_name

    resolved_columns = {}
    for required_column in STAGE_IMPORT_REQUIRED_COLUMNS:
        aliases = STAGE_IMPORT_COLUMN_ALIASES.get(required_column, (required_column,))
        matched_source = None
        for alias in aliases:
            if alias in field_map:
                matched_source = field_map[alias]
                break
        if matched_source is not None:
            resolved_columns[required_column] = matched_source

    missing_columns = [column for column in STAGE_IMPORT_REQUIRED_COLUMNS if column not in resolved_columns]
    if missing_columns:
        return None, [f"В файле нет обязательных колонок: {', '.join(missing_columns)}."]

    normalized_rows = []
    for row_number, raw_row in raw_rows:
        normalized_rows.append(
            {
                "row_number": row_number,
                "order_num": raw_row.get(resolved_columns["order_num"]),
                "title": raw_row.get(resolved_columns["title"]),
                "description": raw_row.get(resolved_columns["description"]),
                "deadline": raw_row.get(resolved_columns["deadline"]),
            }
        )
    return normalized_rows, []


def parse_stage_import_file(uploaded_file):
    # Читает XLSX и отдает строки в единой структуре для валидации.
    extension = Path(uploaded_file.filename or "").suffix.lower()
    if extension != ".xlsx":
        return None, ["Поддерживаются только файлы XLSX."]

    workbook = load_workbook(filename=BytesIO(uploaded_file.read()), data_only=True)
    worksheet = workbook.active
    header = [
        str(cell).strip() if cell is not None else ""
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    ]
    if not any(header):
        return None, ["XLSX-файл не содержит заголовков."]

    raw_rows = []
    for row_number, row_values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = {header[index]: row_values[index] if index < len(row_values) else None for index in range(len(header))}
        if all(str(value or "").strip() == "" for value in row_dict.values()):
            continue
        raw_rows.append((row_number, row_dict))
    return normalize_stage_import_rows(raw_rows, header)


def validate_stage_import_rows(project, parsed_rows):
    # Проверяет строки импорта; при ошибках импорт отменяется целиком.
    errors = []
    prepared_rows = []
    seen_order_nums = set()
    seen_titles = set()

    existing_order_nums = {stage.order_num for stage in project.stages}
    existing_titles = {stage.title.strip().lower() for stage in project.stages}

    for row in parsed_rows:
        row_number = row["row_number"]
        title = str(row["title"] or "").strip()
        description = str(row["description"] or "").strip()

        order_num, order_error = parse_order_num_value(row["order_num"])
        if order_error:
            errors.append(f"Строка {row_number}: {order_error}.")

        deadline, deadline_error = parse_deadline_value(row["deadline"])
        if deadline_error:
            errors.append(f"Строка {row_number}: {deadline_error}.")

        if not title:
            errors.append(f"Строка {row_number}: пустое название этапа.")
        if not description:
            errors.append(f"Строка {row_number}: пустое описание этапа.")

        title_key = title.lower()
        if title and title_key in seen_titles:
            errors.append(f"Строка {row_number}: дубликат названия этапа в файле.")
        if title:
            seen_titles.add(title_key)

        if order_num is not None and order_num in seen_order_nums:
            errors.append(f"Строка {row_number}: дубликат номера этапа в файле.")
        if order_num is not None:
            seen_order_nums.add(order_num)

        if order_num is not None and order_num in existing_order_nums:
            errors.append(f"Строка {row_number}: этап с номером {order_num} уже существует в проекте.")
        if title and title_key in existing_titles:
            errors.append(f"Строка {row_number}: этап с названием '{title}' уже существует в проекте.")

        if not order_error and not deadline_error and title and description:
            prepared_rows.append(
                {
                    "order_num": order_num,
                    "title": title,
                    "description": description,
                    "deadline": deadline,
                }
            )

    return prepared_rows, errors


def has_stage_order_conflict(project_id, order_num, exclude_stage_id=None):
    # Проверка дубликата порядкового номера этапа в проекте.
    query = db.select(Stage.id).filter(
        Stage.project_id == project_id,
        Stage.order_num == order_num,
    )
    if exclude_stage_id is not None:
        query = query.filter(Stage.id != exclude_stage_id)
    existing_id = db.session.execute(query).scalar_one_or_none()
    return existing_id is not None


def build_stage_import_template_workbook():
    # Шаблон импорта этапов: заголовки, пример строки, формат даты.
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Шаблон этапов"
    worksheet.append(list(STAGE_IMPORT_TEMPLATE_HEADERS_RU))
    worksheet.append([1, "Создание команды", "Описание этапа", datetime(2026, 5, 10, 23, 59)])
    worksheet["D2"].number_format = "DD.MM.YYYY HH:MM"
    style_header_row(worksheet)
    widths = [14, 28, 44, 22]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
    worksheet.auto_filter.ref = "A1:D1"
    for col_idx in range(5, 101):
        worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].hidden = True
    worksheet.freeze_panes = "A2"
    return workbook


@main_bp.get("/")
def index():
    # Показывает лендинг и перенаправляет авторизованного пользователя в кабинет.
    if current_user.is_authenticated:
        return redirect(url_for("main.dashboard"))
    return render_template("index.html", hide_guest_nav_actions=True)


@main_bp.get("/health")
def health():
    # Отдает статус приложения для внешней проверки доступности.
    return jsonify({"status": "ok", "service": "ProjectTracker"})


@main_bp.route("/register", methods=["GET", "POST"])
def register():
    # Создает нового пользователя и сразу авторизует после успешной регистрации.
    if current_user.is_authenticated:
        return redirect(url_for("main.dashboard"))

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        password = request.form.get("password", "")
        role = request.form.get("role", "")

        if not email or not first_name or not last_name or not password or role not in {"teacher", "student"}:
            flash("Заполните все поля и выберите корректную роль.")
            return render_template("register.html"), 400

        existing_user = db.session.execute(
            db.select(User).filter_by(email=email)
        ).scalar_one_or_none()
        if existing_user is not None:
            flash("Пользователь с таким email уже существует.")
            return render_template("register.html"), 400

        user = User(
            email=email,
            name=f"{first_name} {last_name}",
            first_name=first_name,
            last_name=last_name,
            password_hash=generate_password_hash(password),
            role=role,
        )
        db.session.add(user)
        db.session.commit()
        login_user(user)
        return redirect(url_for("main.dashboard"))

    return render_template("register.html")


@main_bp.route("/login", methods=["GET", "POST"])
def login():
    # Проверяет учетные данные и открывает пользовательскую сессию.
    if current_user.is_authenticated:
        return redirect(url_for("main.dashboard"))

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        user = db.session.execute(
            db.select(User).filter_by(email=email)
        ).scalar_one_or_none()
        if user is None or not check_password_hash(user.password_hash, password):
            flash("Неверный email или пароль.")
            return render_template("login.html"), 400

        login_user(user)
        return redirect(url_for("main.dashboard"))

    return render_template("login.html")


@main_bp.get("/logout")
@login_required
def logout():
    # Закрывает текущую сессию и возвращает на главную.
    logout_user()
    return redirect(url_for("main.index"))


@main_bp.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    # Показывает и обновляет персональные данные пользователя.
    user = db.session.get(User, current_user.id)
    if user is None:
        abort(404)

    if request.method == "POST":
        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        email = request.form.get("email", "").strip().lower()
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        new_password_confirm = request.form.get("new_password_confirm", "")

        if not first_name or not last_name or not email:
            flash("Имя, фамилия и email обязательны.")
            return render_template("profile.html", user=user), 400

        existing_user = db.session.execute(
            db.select(User).filter(User.email == email, User.id != user.id)
        ).scalar_one_or_none()
        if existing_user is not None:
            flash("Пользователь с таким email уже существует.")
            return render_template("profile.html", user=user), 400

        password_update_requested = any([current_password, new_password, new_password_confirm])
        if password_update_requested:
            if not current_password or not new_password or not new_password_confirm:
                flash("Для смены пароля заполните все поля блока смены пароля.")
                return render_template("profile.html", user=user), 400

            if not check_password_hash(user.password_hash, current_password):
                flash("Текущий пароль указан неверно.")
                return render_template("profile.html", user=user), 400

            if new_password != new_password_confirm:
                flash("Новый пароль и подтверждение не совпадают.")
                return render_template("profile.html", user=user), 400

            user.password_hash = generate_password_hash(new_password)

        user.first_name = first_name
        user.last_name = last_name
        user.email = email
        user.name = f"{first_name} {last_name}"

        log_action("profile_updated", "Обновлены данные профиля пользователя.", user_id=user.id)
        db.session.commit()
        flash("Профиль обновлен.")
        return redirect(url_for("main.profile"))

    return render_template("profile.html", user=user)


@main_bp.get("/dashboard")
@login_required
def dashboard():
    # Разводит пользователя в кабинет учителя или ученика.
    if current_user.role == "teacher":
        return redirect(url_for("main.teacher_dashboard"))
    if current_user.role == "student":
        return redirect(url_for("main.student_dashboard"))
    abort(403)


@main_bp.get("/teacher")
@login_required
@role_required("teacher")
def teacher_dashboard():
    # Точка входа учителя, ведущая к списку проектов.
    return redirect(url_for("main.teacher_projects"))


@main_bp.get("/student")
@login_required
@role_required("student")
def student_dashboard():
    # Точка входа ученика, ведущая к списку проектов.
    return redirect(url_for("main.student_projects"))


@main_bp.get("/teacher/projects")
@login_required
@role_required("teacher")
def teacher_projects():
    # Отображает все проекты текущего учителя.
    projects = db.session.execute(
        db.select(Project).filter_by(teacher_id=current_user.id).order_by(Project.created_at.desc())
    ).scalars()
    return render_template("teacher_projects.html", projects=projects)


@main_bp.get("/teacher/students")
@login_required
@role_required("teacher")
def teacher_students():
    # Показывает каталог учеников и текущие привязки к проекту.
    projects = list(
        db.session.execute(
            db.select(Project).filter_by(teacher_id=current_user.id).order_by(Project.created_at.desc())
        ).scalars()
    )
    students = list(
        db.session.execute(
            db.select(User)
            .filter_by(role="student")
            .order_by(User.last_name.asc(), User.first_name.asc(), User.email.asc())
        ).scalars()
    )

    project_ids = {project.id for project in projects}
    selected_project_id = request.args.get("project_id", type=int)
    if selected_project_id not in project_ids:
        selected_project_id = projects[0].id if projects else None

    linked_student_ids = set()
    if selected_project_id is not None:
        linked_student_ids = {
            link.student_id
            for link in db.session.execute(
                db.select(ProjectStudent).filter_by(project_id=selected_project_id)
            ).scalars()
        }

    return render_template(
        "teacher_students.html",
        projects=projects,
        students=students,
        selected_project_id=selected_project_id,
        linked_student_ids=linked_student_ids,
    )


@main_bp.route("/teacher/project/new", methods=["GET", "POST"])
@login_required
@role_required("teacher")
def teacher_project_new():
    # Создает новый проект учителя.
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()

        if not title or not description:
            flash("Заполните название и описание проекта.")
            return render_template("teacher_project_form.html"), 400

        project = Project(
            title=title,
            description=description,
            teacher_id=current_user.id,
        )
        db.session.add(project)
        db.session.flush()
        log_action(
            "project_created",
            f"Создан проект '{project.title}'",
            project_id=project.id,
        )
        db.session.commit()
        return redirect(url_for("main.teacher_project_detail", project_id=project.id))

    return render_template(
        "teacher_project_form.html",
        project=None,
        form_title="Создать проект",
        submit_label="Сохранить проект",
    )


@main_bp.route("/teacher/project/<int:project_id>/edit", methods=["GET", "POST"])
@login_required
@role_required("teacher")
def teacher_project_edit(project_id):
    # Обновляет название и описание выбранного проекта.
    project = get_teacher_project_or_404(project_id)

    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()

        if not title or not description:
            flash("Заполните название и описание проекта.")
            return render_template(
                "teacher_project_form.html",
                project=project,
                form_title="Редактировать проект",
                submit_label="Сохранить изменения",
            ), 400

        project.title = title
        project.description = description
        log_action(
            "project_updated",
            f"Обновлен проект '{project.title}'",
            project_id=project.id,
        )
        db.session.commit()
        flash("Проект обновлен.")
        return redirect(url_for("main.teacher_project_detail", project_id=project.id))

    return render_template(
        "teacher_project_form.html",
        project=project,
        form_title="Редактировать проект",
        submit_label="Сохранить изменения",
    )


@main_bp.get("/teacher/project/<int:project_id>")
@login_required
@role_required("teacher")
def teacher_project_detail(project_id):
    # Показывает карточку проекта с этапами, учениками и кнопками действий.
    project = get_teacher_project_or_404(project_id)
    unread_by_student = get_teacher_unread_by_student(project.id, current_user.id)
    total_unread = sum(unread_by_student.values())
    return render_template(
        "teacher_project_detail.html",
        project=project,
        unread_by_student=unread_by_student,
        total_unread=total_unread,
    )


@main_bp.post("/teacher/project/<int:project_id>/stages/import")
@login_required
@role_required("teacher")
def teacher_project_import_stages(project_id):
    # Загружает этапы из XLSX и применяет их после полной валидации файла.
    project = get_teacher_project_or_404(project_id)
    uploaded_file = request.files.get("stages_file")

    if uploaded_file is None or not (uploaded_file.filename or "").strip():
        flash("Выберите XLSX файл для импорта этапов.")
        return redirect(url_for("main.teacher_project_detail", project_id=project.id))

    parsed_rows, parse_errors = parse_stage_import_file(uploaded_file)
    if parse_errors:
        for error in parse_errors:
            flash(error)
        return redirect(url_for("main.teacher_project_detail", project_id=project.id))
    if not parsed_rows:
        flash("Файл не содержит строк для импорта.")
        return redirect(url_for("main.teacher_project_detail", project_id=project.id))

    prepared_rows, validation_errors = validate_stage_import_rows(project, parsed_rows)
    if validation_errors:
        flash("Импорт не выполнен: найдены ошибки в файле.")
        for error in validation_errors:
            flash(error)
        return redirect(url_for("main.teacher_project_detail", project_id=project.id))

    stages_to_add = [
        Stage(
            project_id=project.id,
            title=row["title"],
            description=row["description"],
            deadline=row["deadline"],
            order_num=row["order_num"],
        )
        for row in prepared_rows
    ]
    db.session.add_all(stages_to_add)
    log_action(
        "stages_imported",
        f"Импортировано этапов: {len(stages_to_add)}",
        project_id=project.id,
    )
    db.session.commit()
    flash(f"Импорт завершен: добавлено этапов {len(stages_to_add)}.")
    return redirect(url_for("main.teacher_project_detail", project_id=project.id))


@main_bp.get("/teacher/project/<int:project_id>/stages/import/template")
@login_required
@role_required("teacher")
def teacher_project_stage_import_template(project_id):
    # Отдает шаблон XLSX для массового импорта этапов.
    get_teacher_project_or_404(project_id)
    return workbook_response(build_stage_import_template_workbook(), "stage_import_template.xlsx")


@main_bp.get("/teacher/project/<int:project_id>/analytics")
@login_required
@role_required("teacher")
def teacher_project_analytics(project_id):
    # Отображает матрицу прогресса по ученикам и этапам.
    project = get_teacher_project_or_404(project_id)
    stages, student_rows = build_project_analytics_rows(project)

    return render_template(
        "teacher_project_analytics.html",
        project=project,
        stages=stages,
        student_rows=student_rows,
    )


@main_bp.get("/teacher/project/<int:project_id>/analytics/export/xlsx")
@login_required
@role_required("teacher")
def teacher_project_export_analytics_xlsx(project_id):
    # Скачивает XLSX-аналитику по выбранному проекту.
    project = get_teacher_project_or_404(project_id)
    filename = f"project_{project.id}_analytics.xlsx"
    return workbook_response(build_project_analytics_workbook(project), filename)


@main_bp.post("/teacher/project/<int:project_id>/students")
@login_required
@role_required("teacher")
def teacher_project_add_student(project_id):
    # Добавляет ученика в проект по email.
    project = get_teacher_project_or_404(project_id)
    student_email = request.form.get("student_email", "").strip().lower()

    if not student_email:
        flash("Введите email ученика.")
        return redirect(url_for("main.teacher_project_detail", project_id=project.id))

    student = db.session.execute(
        db.select(User).filter_by(email=student_email, role="student")
    ).scalar_one_or_none()
    if student is None:
        flash("Ученик с таким email не найден.")
        return redirect(url_for("main.teacher_project_detail", project_id=project.id))

    existing_link = db.session.execute(
        db.select(ProjectStudent).filter_by(project_id=project.id, student_id=student.id)
    ).scalar_one_or_none()
    if existing_link is not None:
        flash("Этот ученик уже добавлен в проект.")
        return redirect(url_for("main.teacher_project_detail", project_id=project.id))

    db.session.add(ProjectStudent(project_id=project.id, student_id=student.id))
    log_action(
        "student_added_to_project",
        f"В проект '{project.title}' добавлен ученик {student.email}",
        project_id=project.id,
    )
    db.session.commit()
    flash("Ученик успешно добавлен в проект.")
    return redirect(url_for("main.teacher_project_detail", project_id=project.id))


@main_bp.post("/teacher/students/assign")
@login_required
@role_required("teacher")
def teacher_students_assign():
    # Привязывает ученика к проекту из общего списка учеников.
    project_id = request.form.get("project_id", type=int)
    student_id = request.form.get("student_id", type=int)

    if not project_id or not student_id:
        flash("Выберите проект и ученика.")
        return redirect(url_for("main.teacher_students"))

    project = get_teacher_project_or_404(project_id)
    student = db.session.get(User, student_id)
    if student is None or student.role != "student":
        flash("Ученик не найден.")
        return redirect(url_for("main.teacher_students", project_id=project.id))

    existing_link = db.session.execute(
        db.select(ProjectStudent).filter_by(project_id=project.id, student_id=student.id)
    ).scalar_one_or_none()
    if existing_link is not None:
        flash("Этот ученик уже добавлен в проект.")
        return redirect(url_for("main.teacher_students", project_id=project.id))

    db.session.add(ProjectStudent(project_id=project.id, student_id=student.id))
    log_action(
        "student_added_to_project",
        f"В проект '{project.title}' добавлен ученик {student.email}",
        project_id=project.id,
    )
    db.session.commit()
    flash("Ученик успешно добавлен в проект.")
    return redirect(url_for("main.teacher_students", project_id=project.id))


@main_bp.route("/teacher/stage/new", methods=["GET", "POST"])
@login_required
@role_required("teacher")
def teacher_stage_new():
    # Создает новый этап в проекте учителя.
    project_id = request.args.get("project_id", type=int)
    if request.method == "POST":
        project_id = request.form.get("project_id", type=int)

    if not project_id:
        abort(404)

    project = get_teacher_project_or_404(project_id)

    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        deadline_raw = request.form.get("deadline", "").strip()
        order_num = request.form.get("order_num", type=int)

        if not title or not description or not deadline_raw or order_num is None:
            flash("Заполните все поля этапа.")
            return render_template("teacher_stage_form.html", project=project, stage=None), 400

        try:
            deadline = datetime.strptime(deadline_raw, "%Y-%m-%dT%H:%M")
        except ValueError:
            flash("Некорректная дата дедлайна.")
            return render_template("teacher_stage_form.html", project=project, stage=None), 400

        if has_stage_order_conflict(project.id, order_num):
            flash(STAGE_ORDER_DUPLICATE_ERROR)
            return render_template("teacher_stage_form.html", project=project, stage=None), 400

        stage = Stage(
            project_id=project.id,
            title=title,
            description=description,
            deadline=deadline,
            order_num=order_num,
        )
        try:
            db.session.add(stage)
            db.session.flush()
            log_action(
                "stage_created",
                f"Создан этап '{stage.title}' для проекта '{project.title}'",
                project_id=project.id,
            )
            db.session.commit()
            return redirect(url_for("main.teacher_project_detail", project_id=project.id))
        except IntegrityError:
            db.session.rollback()
            flash(STAGE_ORDER_DUPLICATE_ERROR)
            return render_template("teacher_stage_form.html", project=project, stage=None), 400

    return render_template("teacher_stage_form.html", project=project, stage=None)


@main_bp.route("/teacher/stage/<int:stage_id>/edit", methods=["GET", "POST"])
@login_required
@role_required("teacher")
def teacher_stage_edit(stage_id):
    # Обновляет параметры существующего этапа проекта.
    stage = get_teacher_stage_or_404(stage_id)
    project = stage.project

    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        deadline_raw = request.form.get("deadline", "").strip()
        order_num = request.form.get("order_num", type=int)

        if not title or not description or not deadline_raw or order_num is None:
            flash("Заполните все поля этапа.")
            return render_template("teacher_stage_form.html", project=project, stage=stage), 400

        try:
            deadline = datetime.strptime(deadline_raw, "%Y-%m-%dT%H:%M")
        except ValueError:
            flash("Некорректная дата дедлайна.")
            return render_template("teacher_stage_form.html", project=project, stage=stage), 400

        if has_stage_order_conflict(project.id, order_num, exclude_stage_id=stage.id):
            flash(STAGE_ORDER_DUPLICATE_ERROR)
            return render_template("teacher_stage_form.html", project=project, stage=stage), 400

        stage.title = title
        stage.description = description
        stage.deadline = deadline
        stage.order_num = order_num
        try:
            log_action(
                "stage_updated",
                f"Обновлен этап '{stage.title}' проекта '{project.title}'",
                project_id=project.id,
            )
            db.session.commit()
            return redirect(url_for("main.teacher_project_detail", project_id=project.id))
        except IntegrityError:
            db.session.rollback()
            flash(STAGE_ORDER_DUPLICATE_ERROR)
            return render_template("teacher_stage_form.html", project=project, stage=stage), 400

    return render_template("teacher_stage_form.html", project=project, stage=stage)


@main_bp.post("/teacher/stage/<int:stage_id>/delete")
@login_required
@role_required("teacher")
def teacher_stage_delete(stage_id):
    # Удаляет этап из проекта и фиксирует действие в журнале.
    stage = get_teacher_stage_or_404(stage_id)
    project = stage.project
    stage_title = stage.title

    db.session.delete(stage)
    log_action(
        "stage_deleted",
        f"Удален этап '{stage_title}' из проекта '{project.title}'",
        project_id=project.id,
    )
    db.session.commit()
    flash("Этап удален.")
    return redirect(url_for("main.teacher_project_detail", project_id=project.id))


@main_bp.get("/teacher/student/<int:student_id>")
@login_required
@role_required("teacher")
def teacher_student_detail(student_id):
    # Показывает прогресс выбранного ученика в рамках одного проекта.
    project_id = request.args.get("project_id", type=int)
    if not project_id:
        abort(404)

    project = get_teacher_project_or_404(project_id)
    student = db.session.get(User, student_id)
    if student is None or student.role != "student":
        abort(404)

    progress_entries = db.session.execute(
        db.select(Progress)
        .join(Stage)
        .filter(Progress.student_id == student.id, Stage.project_id == project.id)
    ).scalars()
    # Показываем все этапы, даже если записи прогресса еще нет.
    progress_map = {entry.stage_id: entry for entry in progress_entries}

    stage_rows = []
    for stage in project.stages:
        entry = progress_map.get(stage.id)
        status = entry.status if entry else "not_started"
        stage_rows.append(
            {
                "stage": stage,
                "status": status,
                "status_label": get_status_label(status),
                "github_url": entry.github_url if entry else None,
                "comment": entry.comment if entry else None,
            }
        )

    return render_template(
        "teacher_student_detail.html",
        project=project,
        student=student,
        stage_rows=stage_rows,
    )


@main_bp.get("/teacher/project/<int:project_id>/export/xlsx")
@login_required
@role_required("teacher")
def teacher_project_export_xlsx(project_id):
    # Скачивает детальный XLSX-отчет по проекту.
    project = get_teacher_project_or_404(project_id)
    filename = f"project_{project.id}_report.xlsx"
    return workbook_response(build_project_workbook(project), filename)


@main_bp.get("/teacher/project/<int:project_id>/student/<int:student_id>/stage/<int:stage_id>/file")
@login_required
@role_required("teacher")
def teacher_stage_file_download(project_id, student_id, stage_id):
    # Выдает вложение этапа после проверки доступа и безопасного пути.
    project = get_teacher_project_or_404(project_id)
    progress = db.session.execute(
        db.select(Progress)
        .join(Stage)
        .filter(
            Progress.student_id == student_id,
            Progress.stage_id == stage_id,
            Stage.project_id == project.id,
        )
    ).scalar_one_or_none()

    if progress is None or not progress.file_path:
        abort(404)

    # Блокируем доступ к файлам за пределами каталога uploads.
    file_path = Path(progress.file_path).resolve()
    uploads_root = Path(UPLOADS_DIR).resolve()
    if uploads_root not in file_path.parents or not file_path.is_file():
        abort(404)

    return send_file(file_path, as_attachment=True, download_name=file_path.name)


@main_bp.route("/messages/<int:project_id>", methods=["GET", "POST"])
@login_required
def project_messages(project_id):
    # Отображает диалог по проекту и обрабатывает отправку сообщения.
    project = get_project_for_current_user_or_404(project_id)
    participants = []
    unread_by_partner = {}
    selected_partner_id = None

    # В учительском режиме список чатов по ученикам; у ученика только учитель.
    if current_user.role == "teacher":
        participants = [link.student for link in project.student_links]
        participant_ids = {participant.id for participant in participants}
        selected_partner_id = request.args.get("student_id", type=int)
        if selected_partner_id not in participant_ids:
            selected_partner_id = participants[0].id if participants else None
        unread_by_partner = get_teacher_unread_by_student(project.id, current_user.id)
    else:
        participants = [project.teacher]
        selected_partner_id = project.teacher_id
        unread_from_teacher = db.session.execute(
            db.select(db.func.count(Message.id)).filter(
                Message.project_id == project.id,
                Message.from_user_id == project.teacher_id,
                Message.to_user_id == current_user.id,
                Message.is_read.is_(False),
            )
        ).scalar_one()
        unread_by_partner = {project.teacher_id: unread_from_teacher}

    if request.method == "POST":
        text = request.form.get("text", "").strip()
        uploaded_image = request.files.get("image")

        # Учитель пишет только ученикам, привязанным к проекту.
        if current_user.role == "teacher":
            allowed_ids = {participant.id for participant in participants}
            selected_partner_id = request.form.get("student_id", type=int)
            if selected_partner_id not in allowed_ids:
                abort(403)

        # Сообщение может содержать текст, картинку или оба варианта.
        image_path, image_error = validate_and_save_message_image(uploaded_image)
        if image_error:
            flash(image_error)
            if current_user.role == "teacher" and selected_partner_id:
                return redirect(
                    url_for("main.project_messages", project_id=project.id, student_id=selected_partner_id)
                )
            return redirect(url_for("main.project_messages", project_id=project.id))

        if not text and image_path is None:
            flash("Добавьте текст или изображение.")
            if current_user.role == "teacher" and selected_partner_id:
                return redirect(
                    url_for("main.project_messages", project_id=project.id, student_id=selected_partner_id)
                )
            return redirect(url_for("main.project_messages", project_id=project.id))

        recipient_id = selected_partner_id if current_user.role == "teacher" else project.teacher_id
        message = Message(
            project_id=project.id,
            from_user_id=current_user.id,
            to_user_id=recipient_id,
            text=text or "",
            image_path=image_path,
        )
        db.session.add(message)
        db.session.flush()
        recipient = db.session.get(User, recipient_id)
        log_action(
            "message_sent",
            f"Отправлено сообщение пользователю {recipient.email}",
            project_id=project.id,
        )
        db.session.commit()
        if current_user.role == "teacher":
            return redirect(url_for("main.project_messages", project_id=project.id, student_id=recipient_id))
        return redirect(url_for("main.project_messages", project_id=project.id))

    active_partner = db.session.get(User, selected_partner_id) if selected_partner_id else None
    dialog_messages = []
    if selected_partner_id is not None:
        dialog_messages = list(
            db.session.execute(
                db.select(Message)
                .filter(
                    Message.project_id == project.id,
                    db.or_(
                        db.and_(
                            Message.from_user_id == current_user.id,
                            Message.to_user_id == selected_partner_id,
                        ),
                        db.and_(
                            Message.from_user_id == selected_partner_id,
                            Message.to_user_id == current_user.id,
                        ),
                    ),
                )
                .order_by(Message.created_at.asc(), Message.id.asc())
            ).scalars()
        )

    # Прочитанными помечаются только сообщения из открытого диалога.
    has_unread_in_dialog = False
    for message in dialog_messages:
        if (
            message.from_user_id == selected_partner_id
            and message.to_user_id == current_user.id
            and not message.is_read
        ):
            message.is_read = True
            has_unread_in_dialog = True
    if has_unread_in_dialog:
        db.session.commit()
        if current_user.role == "teacher":
            unread_by_partner = get_teacher_unread_by_student(project.id, current_user.id)
        else:
            unread_by_partner = {project.teacher_id: 0}

    user_ids = {current_user.id, project.teacher_id, *(link.student_id for link in project.student_links)}
    users = db.session.execute(db.select(User).filter(User.id.in_(user_ids))).scalars()
    users_map = {user.id: user for user in users}

    return render_template(
        "messages.html",
        project=project,
        participants=participants,
        active_partner=active_partner,
        active_partner_id=selected_partner_id,
        dialog_messages=dialog_messages,
        unread_by_partner=unread_by_partner,
        users_map=users_map,
    )


@main_bp.get("/messages/image/<int:message_id>")
@login_required
def message_image(message_id):
    # Отдает изображение сообщения участнику соответствующего проекта.
    message = db.session.get(Message, message_id)
    if message is None or not message.image_path:
        abort(404)

    get_project_for_current_user_or_404(message.project_id)

    # Картинку сообщения может открыть только участник проекта.
    image_path = Path(message.image_path).resolve()
    uploads_root = Path(UPLOADS_DIR).resolve()
    if uploads_root not in image_path.parents or not image_path.is_file():
        abort(404)

    return send_file(image_path)


@main_bp.get("/api/projects")
@login_required
def api_projects():
    # Возвращает список проектов, доступных текущему пользователю.
    if current_user.role == "teacher":
        projects = db.session.execute(
            db.select(Project).filter_by(teacher_id=current_user.id).order_by(Project.created_at.desc())
        ).scalars()
    else:
        projects = db.session.execute(
            db.select(Project)
            .join(ProjectStudent)
            .filter(ProjectStudent.student_id == current_user.id)
            .order_by(Project.created_at.desc())
        ).scalars()

    payload = [serialize_project(project) for project in projects]
    log_api_call()
    db.session.commit()
    return jsonify(payload)


@main_bp.get("/api/projects/<int:project_id>/students")
@login_required
@role_required("teacher")
def api_project_students(project_id):
    # Возвращает список учеников проекта для его учителя.
    project = get_teacher_project_or_404(project_id)
    students = [
        {
            "id": link.student.id,
            "name": link.student.full_name,
            "email": link.student.email,
        }
        for link in project.student_links
    ]
    log_api_call()
    db.session.commit()
    return jsonify(students)


@main_bp.get("/api/projects/<int:project_id>/stages")
@login_required
def api_project_stages(project_id):
    # Возвращает этапы проекта для его участника.
    project = get_project_for_current_user_or_404(project_id)
    payload = [serialize_stage(stage) for stage in project.stages]
    log_api_call()
    db.session.commit()
    return jsonify(payload)


@main_bp.get("/api/projects/<int:project_id>/stats")
@login_required
@role_required("teacher")
def api_project_stats(project_id):
    # Возвращает сводную статистику статусов по проекту.
    project = get_teacher_project_or_404(project_id)

    entries = db.session.execute(
        db.select(Progress)
        .join(Stage)
        .filter(Stage.project_id == project.id)
    ).scalars()

    stats = {
        "project_id": project.id,
        "total_progress_entries": 0,
        "not_started": 0,
        "in_progress": 0,
        "done": 0,
        "approved": 0,
    }

    for entry in entries:
        stats["total_progress_entries"] += 1
        if entry.status in stats:
            stats[entry.status] += 1

    log_api_call()
    db.session.commit()
    return jsonify(stats)


@main_bp.get("/api/students/<int:student_id>/progress")
@login_required
def api_student_progress(student_id):
    # Возвращает прогресс ученика с учетом прав доступа.
    # Ученик видит только свой прогресс, учитель — только своих учеников.
    if current_user.role == "student" and current_user.id != student_id:
        abort(403)

    student = db.session.get(User, student_id)
    if student is None or student.role != "student":
        abort(404)

    query = db.select(Progress).filter(Progress.student_id == student_id).join(Stage).order_by(Stage.order_num)

    if current_user.role == "teacher":
        query = query.join(Project).filter(Project.teacher_id == current_user.id)

    progress_entries = db.session.execute(query).scalars()
    payload = {
        "student": {
            "id": student.id,
            "name": student.full_name,
            "email": student.email,
        },
        "progress": [
            {
                "stage_id": entry.stage_id,
                "stage_title": entry.stage.title,
                "project_id": entry.stage.project_id,
                "status": entry.status,
                "github_url": entry.github_url,
                "file_path": entry.file_path,
                "comment": entry.comment,
                "updated_at": entry.updated_at.isoformat() if entry.updated_at else None,
            }
            for entry in progress_entries
        ],
    }
    log_api_call()
    db.session.commit()
    return jsonify(payload)


@main_bp.post("/api/messages")
@login_required
def api_messages():
    # Принимает и сохраняет сообщение через REST API.
    data = request.get_json(silent=True) or {}
    project_id = data.get("project_id")
    text = (data.get("text") or "").strip()

    if not project_id or not text:
        return jsonify({"error": "project_id and text are required"}), 400

    project = get_project_for_current_user_or_404(project_id)

    # В API учитель указывает ученика-получателя, ученик пишет учителю.
    if current_user.role == "teacher":
        recipient_id = data.get("recipient_id")
        allowed_ids = {link.student_id for link in project.student_links}
        if recipient_id not in allowed_ids:
            abort(403)
    else:
        recipient_id = project.teacher_id

    message = Message(
        project_id=project.id,
        from_user_id=current_user.id,
        to_user_id=recipient_id,
        text=text,
    )
    db.session.add(message)
    db.session.flush()
    recipient = db.session.get(User, recipient_id)
    log_action(
        "api_message_sent",
        f"Через API отправлено сообщение пользователю {recipient.email}",
        project_id=project.id,
    )
    log_api_call()
    db.session.commit()
    return jsonify({"id": message.id, "status": "created"}), 201


@main_bp.post("/api/progress")
@login_required
@role_required("student")
def api_progress():
    # Обновляет статус этапа ученика через REST API.
    data = request.get_json(silent=True) or {}
    stage_id = data.get("stage_id")
    status = (data.get("status") or "").strip()

    # Через API принимаются только статусы not_started, in_progress, done.
    if not stage_id or status not in {"not_started", "in_progress", "done"}:
        return jsonify({"error": "stage_id and valid status are required"}), 400

    stage = get_student_stage_or_404(stage_id)
    progress = get_or_create_progress(current_user.id, stage.id)
    progress.status = status
    progress.github_url = (data.get("github_url") or "").strip() or None
    progress.comment = (data.get("comment") or "").strip() or None

    log_action(
        "api_progress_updated",
        f"Через API обновлен прогресс по этапу '{stage.title}'",
        project_id=stage.project_id,
    )
    log_api_call()
    db.session.commit()
    return jsonify({"id": progress.id, "status": progress.status}), 201


@main_bp.get("/student/projects")
@login_required
@role_required("student")
def student_projects():
    # Показывает проекты, к которым привязан текущий ученик.
    links = db.session.execute(
        db.select(ProjectStudent)
        .filter_by(student_id=current_user.id)
        .join(Project)
        .order_by(Project.created_at.desc())
    ).scalars()
    return render_template("student_projects.html", links=links)


@main_bp.get("/student/project/<int:project_id>")
@login_required
@role_required("student")
def student_project_detail(project_id):
    # Показывает карточку проекта ученика и список этапов.
    project = get_student_project_or_404(project_id)
    return render_template("student_project_detail.html", project=project)


@main_bp.route("/student/stage/<int:stage_id>", methods=["GET", "POST"])
@login_required
@role_required("student")
def student_stage_detail(stage_id):
    # Показывает этап и сохраняет данные прогресса ученика.
    stage = get_student_stage_or_404(stage_id)

    if request.method == "POST":
        status = request.form.get("status", "").strip()
        github_url = request.form.get("github_url", "").strip()
        comment = request.form.get("comment", "").strip()

        if status not in {"not_started", "in_progress", "done"}:
            flash("Выберите корректный статус.")
            progress = get_or_create_progress(current_user.id, stage.id)
            return render_template("student_stage_detail.html", stage=stage, progress=progress), 400

        progress = get_or_create_progress(current_user.id, stage.id)
        progress.status = status
        progress.github_url = github_url or None
        progress.comment = comment or None
        # Файлы в карточке этапа отключены, остается статус/ссылка/комментарий.
        progress.file_path = None

        log_action(
            "progress_updated",
            f"Обновлен прогресс по этапу '{stage.title}'",
            project_id=stage.project_id,
        )
        db.session.commit()
        flash("Прогресс сохранен.")
        return redirect(url_for("main.student_stage_detail", stage_id=stage.id))

    progress = get_or_create_progress(current_user.id, stage.id)
    db.session.commit()
    return render_template("student_stage_detail.html", stage=stage, progress=progress)
